"""
ARGOS-2 Tool — Document Parser (PDF, CSV, JSON, Excel, Images).

Extracts text content from structured documents.
Supports reading and analyzing file contents across multiple formats.

Dependencies:
  - pypdf (PDF parsing, pure Python, zero system deps)
  - openpyxl (Excel .xlsx parsing)
  - Built-in csv, json modules
  - Vision backend (for analyze_image)
"""

import csv
import json
import os

from .helpers import _get_arg, _normalize_path


def read_pdf_tool(inp):
    """
    Extracts text from a PDF file.

    Input:
        {"filename": "report.pdf"}
        {"filename": "report.pdf", "pages": "1-3"}

    Returns:
        Extracted text content.
    """
    fname = _get_arg(inp, ["filename", "path", "file", "pdf"])
    if not fname:
        return "Error: No filename specified. Use {'filename': 'path/to/file.pdf'}."

    try:
        path = _normalize_path(fname)
    except ValueError as e:
        return f"Error: {e}"
    if not os.path.exists(path):
        return f"Error: File '{path}' not found."
    if not path.lower().endswith(".pdf"):
        return f"Error: '{os.path.basename(path)}' is not a PDF file."

    try:
        from pypdf import PdfReader
    except ImportError:
        return "Error: pypdf not installed. Run: pip install pypdf"

    try:
        reader = PdfReader(path)
        total_pages = len(reader.pages)

        # Parse page range if specified
        page_range = None
        if isinstance(inp, dict) and "pages" in inp:
            page_spec = str(inp["pages"])
            try:
                if "-" in page_spec:
                    start, end = page_spec.split("-", 1)
                    page_range = range(max(0, int(start) - 1), min(total_pages, int(end)))
                else:
                    page_idx = int(page_spec) - 1
                    if 0 <= page_idx < total_pages:
                        page_range = range(page_idx, page_idx + 1)
            except ValueError:
                return f"Error: Invalid page specification '{page_spec}'. Use a number (e.g. '3') or a range (e.g. '1-5')."

        if page_range is None:
            page_range = range(min(total_pages, 20))  # Default: first 20 pages

        text_parts = []
        for i in page_range:
            page_text = reader.pages[i].extract_text()
            if page_text and page_text.strip():
                text_parts.append(f"--- Page {i + 1} ---\n{page_text.strip()}")

        if not text_parts:
            return f"📄 PDF '{os.path.basename(path)}' ({total_pages} pages): No extractable text (may be image-based, use OCR)."

        content = "\n\n".join(text_parts)
        # Truncate very long documents
        if len(content) > 8000:
            content = (
                content[:8000]
                + f"\n\n... [truncated, showing {len(page_range)}/{total_pages} pages]"
            )

        return f"📄 PDF '{os.path.basename(path)}' ({total_pages} pages):\n\n{content}"

    except Exception as e:
        return f"Error reading PDF: {e}"


def read_csv_tool(inp):
    """
    Reads a CSV file and returns a structured summary.

    Input:
        {"filename": "data.csv"}
        {"filename": "data.csv", "rows": 10}

    Returns:
        Column headers and first N rows.
    """
    fname = _get_arg(inp, ["filename", "path", "file", "csv"])
    if not fname:
        return "Error: No filename specified."

    try:
        path = _normalize_path(fname)
    except ValueError as e:
        return f"Error: {e}"
    if not os.path.exists(path):
        return f"Error: File '{path}' not found."

    max_rows = 20
    if isinstance(inp, dict) and "rows" in inp:
        try:
            max_rows = min(int(inp["rows"]), 100)
        except (ValueError, TypeError):
            pass

    try:
        with open(path, encoding="utf-8", errors="replace") as f:
            # Detect delimiter
            sample = f.read(4096)
            f.seek(0)
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            reader = csv.reader(f, dialect)

            rows = []
            for i, row in enumerate(reader):
                rows.append(row)
                if i >= max_rows:
                    break

        if not rows:
            return f"📊 CSV '{os.path.basename(path)}': Empty file."

        # Format as table
        header = " | ".join(rows[0])
        separator = "-" * len(header)
        data_rows = [" | ".join(r) for r in rows[1:]]

        output = f"📊 CSV '{os.path.basename(path)}' ({len(rows) - 1} rows shown):\n\n"
        output += f"{header}\n{separator}\n"
        output += "\n".join(data_rows)

        return output

    except Exception as e:
        return f"Error reading CSV: {e}"


def read_json_tool(inp):
    """
    Reads and pretty-prints a JSON file.

    Input:
        {"filename": "config.json"}

    Returns:
        Formatted JSON content.
    """
    fname = _get_arg(inp, ["filename", "path", "file"])
    if not fname:
        return "Error: No filename specified."

    try:
        path = _normalize_path(fname)
    except ValueError as e:
        return f"Error: {e}"
    if not os.path.exists(path):
        return f"Error: File '{path}' not found."

    try:
        with open(path, encoding="utf-8") as f:
            data = json.load(f)

        formatted = json.dumps(data, indent=2, ensure_ascii=False)
        if len(formatted) > 5000:
            formatted = formatted[:5000] + "\n... [truncated]"

        return f"📋 JSON '{os.path.basename(path)}':\n\n{formatted}"

    except json.JSONDecodeError as e:
        return f"Error: Invalid JSON file: {e}"
    except Exception as e:
        return f"Error reading JSON: {e}"


def read_excel_tool(inp):
    """
    Reads an Excel (.xlsx / .xls) file and returns a structured summary.

    Input:
        {"filename": "data.xlsx"}
        {"filename": "data.xlsx", "sheet": "Sheet1", "rows": 20}

    Returns:
        Sheet names, column headers, and first N rows of the target sheet.
    """
    fname = _get_arg(inp, ["filename", "path", "file", "excel"])
    if not fname:
        return "Error: No filename specified."

    try:
        path = _normalize_path(fname)
    except ValueError as e:
        return f"Error: {e}"
    if not os.path.exists(path):
        return f"Error: File '{path}' not found."
    if not path.lower().endswith((".xlsx", ".xls", ".xlsm")):
        return f"Error: '{os.path.basename(path)}' is not an Excel file (.xlsx/.xls/.xlsm)."

    try:
        import openpyxl
    except ImportError:
        return "Error: openpyxl not installed. Run: pip install openpyxl"

    max_rows = 20
    if isinstance(inp, dict) and "rows" in inp:
        try:
            max_rows = min(int(inp["rows"]), 100)
        except (ValueError, TypeError):
            pass

    target_sheet = inp.get("sheet") if isinstance(inp, dict) else None

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames

        if target_sheet and target_sheet in sheet_names:
            ws = wb[target_sheet]
        else:
            ws = wb.active
            target_sheet = ws.title

        rows = []
        for row in ws.iter_rows(max_row=max_rows + 1, values_only=True):
            rows.append([str(cell) if cell is not None else "" for cell in row])

        wb.close()

        if not rows:
            return f"📊 Excel '{os.path.basename(path)}': Sheet '{target_sheet}' is empty."

        header = " | ".join(rows[0])
        separator = "-" * min(len(header), 120)
        data_rows = [" | ".join(r) for r in rows[1:]]

        output = (
            f"📊 Excel '{os.path.basename(path)}' — Sheet: '{target_sheet}'\n"
            f"Available sheets: {', '.join(sheet_names)}\n\n"
            f"{header}\n{separator}\n"
        )
        output += "\n".join(data_rows)
        return output

    except Exception as e:
        return f"Error reading Excel file: {e}"


def analyze_image_tool(inp):
    """
    Analyzes an image file using the vision backend.
    Supports PNG, JPEG, GIF, BMP, WEBP, TIFF.

    Input:
        {"filename": "chart.png"}
        {"filename": "chart.png", "question": "What values are shown in this bar chart?"}

    Returns:
        Vision model description / answer to the question.
    """
    fname = _get_arg(inp, ["filename", "path", "file", "image"])
    if not fname:
        return "Error: No filename specified."

    try:
        path = _normalize_path(fname)
    except ValueError as e:
        return f"Error: {e}"

    question = (
        "Describe this image in detail, extracting all visible text, numbers, labels, and data."
    )
    if isinstance(inp, dict) and inp.get("question"):
        question = inp["question"]

    try:
        from src import vision

        return vision.analyze_image_file(path, question)
    except Exception as e:
        return f"Error analyzing image: {e}"


def query_table_tool(inp):
    """
    Runs a pandas query/aggregation on a CSV or Excel file.
    Ideal for tasks that require filtering, grouping, or computing statistics.

    Input:
        {"filename": "data.csv", "filter": "year == 2020 and value > 100"}
        {"filename": "data.xlsx", "group_by": "country", "aggregate": "sum", "select": ["gdp"]}
        {"filename": "data.csv", "filter": "category == 'A'", "aggregate": "mean"}

    Supported aggregates: sum, mean, count, max, min, median, std
    """
    fname = _get_arg(inp, ["filename", "path", "file"])
    if not fname:
        return "Error: No filename specified."

    try:
        path = _normalize_path(fname)
    except ValueError as e:
        return f"Error: {e}"
    if not os.path.exists(path):
        return f"Error: File '{path}' not found."

    try:
        import pandas as pd
    except ImportError:
        return "Error: pandas not installed. Run: pip install pandas"

    try:
        lower = path.lower()
        if lower.endswith((".xlsx", ".xls", ".xlsm")):
            sheet = inp.get("sheet") if isinstance(inp, dict) else None
            df = pd.read_excel(path, sheet_name=sheet or 0)
        elif lower.endswith(".csv"):
            df = pd.read_csv(path)
        else:
            return "Error: Unsupported file type. Use CSV or Excel."
    except Exception as e:
        return f"Error loading file: {e}"

    # Filter rows
    if isinstance(inp, dict) and inp.get("filter"):
        try:
            df = df.query(inp["filter"])
        except Exception as e:
            return f"Error in filter '{inp['filter']}': {e}"

    # Group + aggregate (select is applied to the result, not before, so group_by col is preserved)
    if isinstance(inp, dict) and inp.get("aggregate"):
        agg = inp["aggregate"]
        valid_aggs = {"sum", "mean", "count", "max", "min", "median", "std"}
        if agg not in valid_aggs:
            return f"Error: Invalid aggregate '{agg}'. Use one of: {', '.join(sorted(valid_aggs))}"
        group_by = inp.get("group_by")
        # numeric_only avoids errors on string columns for reduction aggregates
        numeric_aggs = {"sum", "mean", "median", "std"}
        kwargs = {"numeric_only": True} if agg in numeric_aggs else {}
        try:
            if group_by:
                result = getattr(df.groupby(group_by), agg)(**kwargs)
            else:
                result = getattr(df, agg)(**kwargs)
            # Apply column selection to the result if requested
            select = inp.get("select") if isinstance(inp, dict) else None
            if select:
                cols = select if isinstance(select, list) else [select]
                existing = [
                    c
                    for c in cols
                    if c in result.index or (hasattr(result, "columns") and c in result.columns)
                ]
                if existing and hasattr(result, "columns"):
                    result = result[existing]
            return f"Result ({agg}{' by ' + group_by if group_by else ''}):\n{result.to_string()}"
        except Exception as e:
            return f"Error computing {agg}: {e}"

    # Select columns (only when not aggregating)
    if isinstance(inp, dict) and inp.get("select"):
        cols = inp["select"] if isinstance(inp["select"], list) else [inp["select"]]
        missing = [c for c in cols if c not in df.columns]
        if missing:
            return f"Error: Columns not found: {missing}. Available: {df.columns.tolist()}"
        df = df[cols]

    # Default: show head
    rows_to_show = min(len(df), 50)
    output = (
        f"📊 '{os.path.basename(path)}' — {df.shape[0]} rows × {df.shape[1]} cols\n"
        f"Columns: {', '.join(df.columns.tolist())}\n\n"
    )
    output += df.head(rows_to_show).to_string(index=False)
    if len(df) > rows_to_show:
        output += f"\n... [{len(df) - rows_to_show} more rows]"
    return output


def transcribe_audio_tool(inp):
    """
    Transcribes an audio file to text using speech recognition.
    Supports WAV, FLAC, AIFF, OGG formats.

    Input:
        {"filename": "audio.wav"}
        {"filename": "recording.flac", "language": "it-IT"}

    Note: MP3 files must be converted first (e.g. via bash_exec with ffmpeg).
    """
    fname = _get_arg(inp, ["filename", "path", "file", "audio"])
    if not fname:
        return "Error: No filename specified."

    try:
        path = _normalize_path(fname)
    except ValueError as e:
        return f"Error: {e}"
    if not os.path.exists(path):
        return f"Error: File '{path}' not found."

    supported = (".wav", ".flac", ".aiff", ".aif", ".ogg")
    if not path.lower().endswith(supported):
        return (
            f"Error: Unsupported audio format '{os.path.splitext(path)[1]}'. "
            f"Supported: {', '.join(supported)}. "
            "For MP3, convert first: bash_exec with 'ffmpeg -i input.mp3 output.wav'"
        )

    language = "en-US"
    if isinstance(inp, dict) and inp.get("language"):
        language = inp["language"]

    try:
        import speech_recognition as sr
    except ImportError:
        return "Error: SpeechRecognition not installed. Run: pip install SpeechRecognition"

    recognizer = sr.Recognizer()
    try:
        with sr.AudioFile(path) as source:
            audio = recognizer.record(source)
    except Exception as e:
        return f"Error reading audio file: {e}"

    try:
        text = recognizer.recognize_google(audio, language=language)
        return f"🎤 Transcription of '{os.path.basename(path)}':\n\n{text}"
    except sr.UnknownValueError:
        return (
            f"Could not understand audio in '{os.path.basename(path)}' (speech unclear or silent)."
        )
    except sr.RequestError as e:
        return f"Speech recognition service error: {e}"
    except Exception as e:
        return f"Transcription failed: {e}"
